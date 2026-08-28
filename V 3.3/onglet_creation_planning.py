# -*- coding: utf-8 -*-
"""
onglet_creation_planning.py
-----------------------------
Onglet "🌱 Création planning cultural" : assistant pas-à-pas (pop-ups
successives) qui aide l'utilisateur à construire un planning cultural pour
sa ferme, puis rassemble le résultat dans un tableau récapitulatif intégré
à l'onglet et exportable vers le fichier Excel du planning.

Déroulé de l'assistant (voir lancer_assistant) :
    1. Point de départ  : "je commence de zéro" / "j'ai déjà cultivé"
    2. Durée de culture  : période (dates) sur laquelle générer le planning
    3. Région            : carte régionale (climat/sol) - cf. carte_france.py
    4. Disposition       : structure Zones/Chapelles/Planches de la ferme
                            (générée par défaut, ou saisie/importée si
                            l'utilisateur a déjà cultivé - avec la dernière
                            culture de chaque planche)
    5. Sélection cultures : pour chaque planche, choix d'une culture parmi
                             celles triées par score d'adéquation décroissant
                             (région + saison + [rotation])

Ce module est volontairement coupé en deux parties, comme le reste de
l'application :
  - une partie logique pure (scoring, génération de disposition, tableau
    final) qui ne dépend pas de PyQt et peut être testée séparément ;
  - une partie interface (pop-ups QDialog + construction de l'onglet).

Le calcul de score est un système à règles simple et documenté (comme
meteo_decision.py) : il donne un repère de départ raisonné, pas une vérité
absolue - à ajuster selon la connaissance du terrain.
"""

import datetime
import os
import re

from PyQt5.QtCore import Qt
from PyQt5.QtGui import QColor, QBrush
from PyQt5.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QFormLayout, QLabel, QRadioButton,
    QButtonGroup, QDialogButtonBox, QDateEdit, QSpinBox, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView,
    QComboBox, QListWidget, QListWidgetItem, QGroupBox, QMessageBox,
    QWidget, QSplitter, QTextBrowser, QFileDialog, QLineEdit, QGraphicsScene,
)


# ====================================================================
# PARTIE LOGIQUE PURE (pas de dépendance PyQt)
# ====================================================================

# ------------------------------------------------------------------
# Score "région" : compatibilité climat/sol de la région avec la culture
# ------------------------------------------------------------------
_RE_PLAGE_TEMP = re.compile(r"(-?\d+)\s*-\s*(-?\d+)\s*°C")

# Besoin en sol (issu de rotation_cultures.classifier_besoin) -> niveau de
# fertilité de sol idéal (échelle 1-5 de carte_france.SOL_LABELS_NIVEAU).
_SOL_IDEAL_PAR_GROUPE = {
    "Gourmande (très exigeante)": 5,
    "Moyennement exigeante": 4,
    "Légumineuse (fixe l'azote)": 3,
    "Peu exigeante (sol reposé / racine)": 3,
}


def temp_ideale_croissance(fiche):
    """Extrait (tmin, tmax) en °C depuis le champ texte 'temp_croissance'
    d'une fiche botanique (ex. "18-26 °C, nouaison compromise..." -> (18,
    26)). Renvoie None si aucune plage n'a pu être trouvée."""
    m = _RE_PLAGE_TEMP.search(fiche.get("temp_croissance", "") or "")
    if not m:
        return None
    a, b = float(m.group(1)), float(m.group(2))
    return (min(a, b), max(a, b))


def score_region(fiche, nom_region, cf):
    """Score 0-100 d'adéquation climat/sol d'une région pour une culture.

    Heuristique volontairement simple (comme carte_france.py le précise
    lui-même, les données régionales sont indicatives) :
      - température : on compare le milieu de la plage de température de
        croissance idéale de la culture à une estimation grossière de la
        température en pleine saison de la région (température moyenne
        annuelle indicative + 8°C, pour approcher un "cœur d'été/saison") ;
      - sol : on compare le niveau de fertilité indicatif de la région au
        niveau idéal pour le groupe de besoin de la culture (gourmande,
        moyenne, légumineuse, peu exigeante).
    Renvoie un score global (moyenne pondérée 60% température / 40% sol).
    """
    donnees_region = cf.REGIONS[nom_region]

    plage = temp_ideale_croissance(fiche)
    if plage is None:
        score_temp = 70.0  # pas d'info : score neutre, ni pénalisant ni favorisant
    else:
        milieu_culture = (plage[0] + plage[1]) / 2
        temp_saison_region = donnees_region["temp_moy_indicatif_c"] + 8
        ecart = abs(milieu_culture - temp_saison_region)
        score_temp = max(0.0, 100.0 - ecart * 6.0)

    groupe = _groupe_besoin(fiche)
    sol_ideal = _SOL_IDEAL_PAR_GROUPE.get(groupe, 4)
    ecart_sol = abs(donnees_region["sol_niveau"] - sol_ideal)
    score_sol = max(0.0, 100.0 - ecart_sol * 15.0)

    return round(0.6 * score_temp + 0.4 * score_sol)


def _groupe_besoin(fiche):
    """Classement en groupe de besoin (même logique que
    rotation_cultures.classifier_besoin, dupliquée ici pour ne pas imposer
    de dépendance directe au module rc dans le calcul de score région)."""
    texte = (fiche.get("fertilisation") or "").lower()
    if "légumineuses" in texte and "azote" in texte:
        return "Légumineuse (fixe l'azote)"
    if "très gourmande" in texte or "gourmande" in texte:
        return "Gourmande (très exigeante)"
    if ("éviter tout apport de matière organique fraîche" in texte
            or "fourchage" in texte or "à l'automne précédent" in texte
            or "peu d'azote" in texte):
        return "Peu exigeante (sol reposé / racine)"
    return "Moyennement exigeante"


# ------------------------------------------------------------------
# Score "saison" : proximité de la semaine actuelle avec la fenêtre de
# lancement de la culture (semis/plantation) dans le planning existant.
# ------------------------------------------------------------------
ACTIONS_LANCEMENT = ("Semis direct", "Semis en pot/plant", "Plantation")


def _semaine_dans_fenetre(semaine, debut, fin):
    if debut <= fin:
        return debut <= semaine <= fin
    return semaine >= debut or semaine <= fin


def _distance_semaines(semaine, debut):
    return (debut - semaine) % 52


def score_saison(culture_nom, semaine_ref, rows_planning):
    """Score 0-100 selon la proximité de semaine_ref avec la/les fenêtre(s)
    de semis/plantation de cette culture dans le planning existant (CSV).
    Renvoie (score, meilleure_fenetre_texte)."""
    fenetres = [
        (int(r["semaine_debut"]), int(r["semaine_fin"]))
        for r in rows_planning
        if r["culture"].strip().lower() == culture_nom.strip().lower()
        and r["action"] in ACTIONS_LANCEMENT
        and str(r.get("semaine_debut", "")).isdigit()
        and str(r.get("semaine_fin", "")).isdigit()
    ]
    if not fenetres:
        return 50, "fenêtre inconnue"

    meilleur_score = -1
    meilleur_texte = ""
    for debut, fin in fenetres:
        if _semaine_dans_fenetre(semaine_ref, debut, fin):
            score = 100
        else:
            dist = min(_distance_semaines(semaine_ref, debut),
                       _distance_semaines(fin, semaine_ref))
            score = max(0, 100 - dist * 10)
        if score > meilleur_score:
            meilleur_score = score
            meilleur_texte = f"sem. {debut}-{fin}"
    return meilleur_score, meilleur_texte


# ------------------------------------------------------------------
# Score "rotation" : uniquement si la planche a déjà porté une culture
# (règle impérative de non-retour de la même famille botanique).
# ------------------------------------------------------------------
def score_rotation(culture_nom, derniere_culture, fb, rc):
    """Renvoie (score, motif). score=0 signifie "interdit" (même famille
    botanique que la dernière culture de la planche - règle impérative)."""
    if not derniere_culture or derniere_culture not in fb.FICHES:
        return 100, ""
    analyse = rc.analyser_rotation(derniere_culture)
    if analyse is None:
        return 100, ""
    noms_a_eviter = {e["nom"] for e in analyse["a_eviter"]}
    noms_recommandees = {e["nom"] for e in analyse["recommandees"]}
    noms_possibles = {e["nom"] for e in analyse["possibles"]}
    if culture_nom in noms_a_eviter:
        return 0, f"même famille que « {derniere_culture} » (délai conseillé : {analyse['delai_retour_annees']} ans)"
    if culture_nom in noms_recommandees:
        return 100, "rotation recommandée après cette culture"
    if culture_nom in noms_possibles:
        return 70, "rotation possible"
    return 55, ""


# ------------------------------------------------------------------
# Agrégation : liste des cultures triées par score pour une planche donnée
# ------------------------------------------------------------------
def lister_cultures_triees(nom_region, semaine_ref, derniere_culture,
                            rows_planning, fb, rc, cf):
    """Renvoie une liste de dicts {culture, famille, score, score_region,
    score_saison, score_rotation, fenetre, motif_rotation, interdit},
    triée par score décroissant. `derniere_culture` peut être vide/None
    (mode "je commence de zéro" : pas de contrainte de rotation)."""
    resultats = []
    for nom_culture, fiche in fb.FICHES.items():
        s_region = score_region(fiche, nom_region, cf)
        s_saison, fenetre = score_saison(nom_culture, semaine_ref, rows_planning)
        s_rotation, motif = score_rotation(nom_culture, derniere_culture, fb, rc)

        if s_rotation == 0:
            score_final = 0
        elif derniere_culture:
            score_final = round(0.4 * s_region + 0.3 * s_saison + 0.3 * s_rotation)
        else:
            score_final = round(0.55 * s_region + 0.45 * s_saison)

        resultats.append({
            "culture": nom_culture,
            "famille": fiche["famille"],
            "score": score_final,
            "score_region": s_region,
            "score_saison": s_saison,
            "score_rotation": s_rotation,
            "fenetre": fenetre,
            "motif_rotation": motif,
            "interdit": s_rotation == 0,
        })

    resultats.sort(key=lambda e: e["score"], reverse=True)
    return resultats


# ------------------------------------------------------------------
# Disposition de la ferme (Zones > Chapelles > Planches)
# ------------------------------------------------------------------
def generer_disposition_defaut(nb_zones, nb_chapelles, nb_planches):
    """Disposition par défaut proposée à un débutant : nb_zones zones,
    chacune avec nb_chapelles chapelles, chacune avec nb_planches planches.
    Renvoie une liste de dicts {zone, chapelle, planche, derniere_culture}
    (derniere_culture toujours vide dans ce cas)."""
    planches = []
    for z in range(1, nb_zones + 1):
        nom_zone = f"Zone {z}"
        for c in range(1, nb_chapelles + 1):
            nom_chapelle = f"Chapelle {z}.{c}"
            for p in range(1, nb_planches + 1):
                planches.append({
                    "zone": nom_zone,
                    "chapelle": nom_chapelle,
                    "planche": f"Planche {z}.{c}.{p}",
                    "derniere_culture": "",
                })
    return planches


def importer_planches_depuis_plan(chemin, pf):
    """Lit un fichier .ferme (format plan_ferme.py) et renvoie la liste des
    planches trouvées, sous la même forme que generer_disposition_defaut,
    en reprenant la culture actuellement affectée à chaque planche (champ
    'culture' du rectangle) comme 'derniere_culture'."""
    with open(chemin, "r", encoding="utf-8") as f:
        rects = pf.deserialiser(f.read())

    par_id = {r["id"]: r for r in rects}

    def nom_ancetre(rect, type_cherche):
        courant = rect
        while courant is not None:
            if courant["type"] == type_cherche:
                return courant["nom"]
            courant = par_id.get(courant["parent"]) if courant["parent"] else None
        return ""

    planches = []
    for r in rects:
        if r["type"] != "Planche":
            continue
        planches.append({
            "zone": nom_ancetre(r, "Zone") or "(sans zone)",
            "chapelle": nom_ancetre(par_id.get(r["parent"], r), "Chapelle") or "(sans chapelle)",
            "planche": r["nom"],
            "derniere_culture": r.get("culture", ""),
        })
    return planches


# ------------------------------------------------------------------
# Tableau final : rassemble, pour chaque planche assignée, les lignes du
# planning cultural existant (CSV) correspondant à la culture choisie et à
# la période demandée.
# ------------------------------------------------------------------
COLONNES_FINALES = ["zone", "chapelle", "planche", "culture", "conduite",
                     "variete_n", "action", "semaine_debut", "semaine_fin",
                     "mois_debut", "mois_fin", "commentaire", "score"]
ENTETES_FINALES = ["Zone", "Chapelle", "Planche", "Culture", "Conduite",
                    "Variété n°", "Action", "Sem. début", "Sem. fin",
                    "Mois début", "Mois fin", "Commentaire", "Score adéquation"]


def construire_lignes_finales(planches, assignations, scores_par_planche,
                               semaine_debut_periode, semaine_fin_periode,
                               rows_planning):
    """planches : liste de dicts (zone/chapelle/planche/derniere_culture).
    assignations : dict {index_planche: nom_culture}.
    scores_par_planche : dict {index_planche: score_int} (score de la
    culture choisie, pour affichage dans le tableau).

    Pour chaque planche assignée, reprend dans rows_planning (le planning
    CSV déjà chargé dans l'application) toutes les lignes de la culture
    choisie dont la fenêtre de semaines chevauche la période demandée, et
    les rattache à cette planche. Si aucune ligne du planning existant ne
    correspond à la culture, une ligne indicative unique est tout de même
    créée (action vide) pour que la planche apparaisse dans le tableau."""
    lignes = []
    for idx, culture in assignations.items():
        planche = planches[idx]
        score = scores_par_planche.get(idx, "")
        lignes_culture = [
            r for r in rows_planning
            if r["culture"].strip().lower() == culture.strip().lower()
            and str(r.get("semaine_debut", "")).isdigit()
            and str(r.get("semaine_fin", "")).isdigit()
            and _periodes_se_chevauchent(
                int(r["semaine_debut"]), int(r["semaine_fin"]),
                semaine_debut_periode, semaine_fin_periode)
        ]
        if not lignes_culture:
            lignes.append({
                "zone": planche["zone"], "chapelle": planche["chapelle"],
                "planche": planche["planche"], "culture": culture,
                "conduite": "", "variete_n": "", "action": "(aucune donnée de planning)",
                "semaine_debut": "", "semaine_fin": "", "mois_debut": "", "mois_fin": "",
                "commentaire": "", "score": score,
            })
            continue
        for r in lignes_culture:
            lignes.append({
                "zone": planche["zone"], "chapelle": planche["chapelle"],
                "planche": planche["planche"], "culture": culture,
                "conduite": r.get("conduite", ""), "variete_n": r.get("variete_n", ""),
                "action": r.get("action", ""), "semaine_debut": r.get("semaine_debut", ""),
                "semaine_fin": r.get("semaine_fin", ""), "mois_debut": r.get("mois_debut", ""),
                "mois_fin": r.get("mois_fin", ""), "commentaire": r.get("commentaire", ""),
                "score": score,
            })
    lignes.sort(key=lambda l: (l["zone"], l["chapelle"], l["planche"]))
    return lignes


def _periodes_se_chevauchent(debut_a, fin_a, debut_b, fin_b):
    """Teste le chevauchement de deux intervalles de semaines (1-52),
    chacun pouvant chevaucher le changement d'année (debut > fin)."""
    semaines_a = _semaines_de_intervalle(debut_a, fin_a)
    semaines_b = _semaines_de_intervalle(debut_b, fin_b)
    return not semaines_a.isdisjoint(semaines_b)


def _semaines_de_intervalle(debut, fin):
    if debut <= fin:
        return set(range(debut, fin + 1))
    return set(range(debut, 53)) | set(range(1, fin + 1))


def exporter_vers_excel(chemin_xlsx, lignes, nom_feuille="Planning créé (assistant)"):
    """Écrit (ou remplace) une feuille dans le classeur Excel existant avec
    les lignes du tableau récapitulatif. Nécessite openpyxl."""
    import openpyxl

    if os.path.exists(chemin_xlsx):
        classeur = openpyxl.load_workbook(chemin_xlsx)
    else:
        classeur = openpyxl.Workbook()
        classeur.remove(classeur.active)

    if nom_feuille in classeur.sheetnames:
        del classeur[nom_feuille]
    feuille = classeur.create_sheet(nom_feuille)

    feuille.append(ENTETES_FINALES)
    for ligne in lignes:
        feuille.append([ligne.get(c, "") for c in COLONNES_FINALES])

    for i, entete in enumerate(ENTETES_FINALES, start=1):
        feuille.column_dimensions[feuille.cell(row=1, column=i).column_letter].width = max(12, len(entete) + 2)

    classeur.save(chemin_xlsx)


# ====================================================================
# PARTIE INTERFACE (PyQt)
# ====================================================================

class DialoguePointDepart(QDialog):
    """Pop-up 1 : point de départ de l'assistant."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Créer un planning cultural — Point de départ")
        self.resize(460, 220)
        layout = QVBoxLayout(self)

        layout.addWidget(QLabel(
            "<b>Avant de commencer, une question :</b>"))

        self.groupe = QButtonGroup(self)
        self.radio_zero = QRadioButton("🌱 Je commence de zéro (nouvelle ferme, planches vierges)")
        self.radio_existant = QRadioButton("🚜 J'ai déjà cultivé sur mes terres (rotation à respecter)")
        self.radio_zero.setChecked(True)
        self.groupe.addButton(self.radio_zero)
        self.groupe.addButton(self.radio_existant)
        layout.addWidget(self.radio_zero)
        layout.addWidget(self.radio_existant)
        layout.addStretch(1)

        info = QLabel(
            "Si vous avez déjà cultivé, l'assistant vous demandera la dernière culture "
            "de chaque planche afin de respecter la règle de rotation (ne jamais "
            "refaire suivre une culture par une autre de la même famille botanique)."
        )
        info.setWordWrap(True)
        info.setStyleSheet("color:#555; font-style:italic; font-size:9pt;")
        layout.addWidget(info)

        boutons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        boutons.accepted.connect(self.accept)
        boutons.rejected.connect(self.reject)
        layout.addWidget(boutons)

    def mode(self):
        return "existant" if self.radio_existant.isChecked() else "zero"


class DialogueDureeCulture(QDialog):
    """Pop-up 2 : période sur laquelle générer le planning."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Créer un planning cultural — Durée de culture")
        self.resize(420, 180)
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(
            "Sur quelle période souhaitez-vous générer le planning ?"))

        form = QFormLayout()
        aujourd_hui = datetime.date.today()
        self.date_debut = QDateEdit(calendarPopup=True)
        self.date_debut.setDate(aujourd_hui)
        self.date_fin = QDateEdit(calendarPopup=True)
        self.date_fin.setDate(aujourd_hui + datetime.timedelta(days=120))
        form.addRow("Début :", self.date_debut)
        form.addRow("Fin :", self.date_fin)
        layout.addLayout(form)
        layout.addStretch(1)

        boutons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        boutons.accepted.connect(self._valider)
        boutons.rejected.connect(self.reject)
        layout.addWidget(boutons)

    def _valider(self):
        if self.date_fin.date() < self.date_debut.date():
            QMessageBox.warning(self, "Dates invalides", "La date de fin doit être postérieure à la date de début.")
            return
        self.accept()

    @staticmethod
    def _semaine(qdate):
        d = datetime.date(qdate.year(), qdate.month(), qdate.day())
        return max(1, min(52, ((d.timetuple().tm_yday - 1) // 7) + 1))

    def periode(self):
        return self._semaine(self.date_debut.date()), self._semaine(self.date_fin.date())


class DialogueChoixRegion(QDialog):
    """Pop-up 3 : sélection de la région (climat/sol) sur une carte
    schématique, en réutilisant les tuiles de carte_france.py."""

    def __init__(self, parent, cf, region_par_defaut=None):
        super().__init__(parent)
        self.cf = cf
        self.setWindowTitle("Créer un planning cultural — Région de la ferme")
        self.resize(760, 520)
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(
            "Cliquez sur votre région : le climat et le sol indicatifs de la région "
            "sont utilisés pour trier les cultures les mieux adaptées."))

        splitter = QSplitter(Qt.Horizontal)
        layout.addWidget(splitter, 1)

        self.scene = QGraphicsScene()
        self.vue = cf.VueCarteFrance(self.scene)
        self.vue.setMinimumWidth(420)
        splitter.addWidget(self.vue)

        self.texte_info = QTextBrowser()
        splitter.addWidget(self.texte_info)
        splitter.setStretchFactor(0, 3)
        splitter.setStretchFactor(1, 2)

        self.tuiles = {}
        for nom, d in self.cf.REGIONS.items():
            cx, cy = self.cf.position_pixel(d["row"], d["col"])
            tuile = self.cf.TuileRegion(nom, cx, cy, callback_clic=self._selectionner)
            tuile.definir_couleur(self.cf.couleur_pour(nom, 0))
            self.scene.addItem(tuile)
            self.scene.addItem(self.cf.creer_etiquette(nom, cx, cy))
            self.tuiles[nom] = tuile
        self.scene.setSceneRect(self.scene.itemsBoundingRect().adjusted(-40, -40, 40, 40))

        self._region_choisie = region_par_defaut if region_par_defaut in self.cf.REGIONS else next(iter(self.cf.REGIONS))
        self._selectionner(self._region_choisie)

        boutons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        boutons.accepted.connect(self.accept)
        boutons.rejected.connect(self.reject)
        layout.addWidget(boutons)

    def showEvent(self, event):
        super().showEvent(event)
        self.vue.ajuster_vue()

    def _selectionner(self, nom):
        self._region_choisie = nom
        for autre, tuile in self.tuiles.items():
            tuile.definir_selection(autre == nom)
        d = self.cf.REGIONS[nom]
        self.texte_info.setHtml(
            f"<h3>{nom}</h3>"
            f"<b>🌦 Climat :</b> {d['climat']}<br>{d['climat_desc']}<br><br>"
            f"<b>🧪 Sol :</b> {d['sol_type']} "
            f"<i>({self.cf.SOL_LABELS_NIVEAU[d['sol_niveau']]})</i><br>{d['sol_desc']}"
        )

    def region_choisie(self):
        return self._region_choisie


class DialogueDispositionFerme(QDialog):
    """Pop-up 4 : disposition Zones/Chapelles/Planches de la ferme."""

    def __init__(self, parent, mode, fb, pf):
        super().__init__(parent)
        self.mode = mode
        self.fb = fb
        self.pf = pf
        self.setWindowTitle("Créer un planning cultural — Disposition de la ferme")
        self.resize(720, 520)
        self._planches = []

        layout = QVBoxLayout(self)

        if mode == "zero":
            layout.addWidget(QLabel(
                "Disposition par défaut proposée pour un débutant : ajustez les nombres "
                "puis cliquez sur « Générer », ou modifiez directement les noms dans le tableau."))
            barre = QHBoxLayout()
            self.spin_zones = QSpinBox(); self.spin_zones.setRange(1, 20); self.spin_zones.setValue(1)
            self.spin_chapelles = QSpinBox(); self.spin_chapelles.setRange(1, 20); self.spin_chapelles.setValue(2)
            self.spin_planches = QSpinBox(); self.spin_planches.setRange(1, 50); self.spin_planches.setValue(4)
            barre.addWidget(QLabel("Zones :")); barre.addWidget(self.spin_zones)
            barre.addWidget(QLabel("Chapelles/zone :")); barre.addWidget(self.spin_chapelles)
            barre.addWidget(QLabel("Planches/chapelle :")); barre.addWidget(self.spin_planches)
            btn_generer = QPushButton("🔄 Générer la disposition par défaut")
            btn_generer.clicked.connect(self._generer_defaut)
            barre.addWidget(btn_generer)
            barre.addStretch(1)
            layout.addLayout(barre)
        else:
            layout.addWidget(QLabel(
                "Renseignez vos planches existantes et, pour chacune, la dernière culture "
                "réalisée (utilisée pour respecter la rotation). Vous pouvez aussi importer "
                "un plan de ferme déjà enregistré (onglet « Plan de la ferme »)."))
            barre = QHBoxLayout()
            btn_ajouter = QPushButton("➕ Ajouter une planche")
            btn_ajouter.clicked.connect(self._ajouter_ligne_vide)
            btn_supprimer = QPushButton("🗑 Supprimer la ligne sélectionnée")
            btn_supprimer.clicked.connect(self._supprimer_ligne)
            btn_importer = QPushButton("📂 Importer un plan de ferme (.ferme)...")
            btn_importer.clicked.connect(self._importer_plan)
            barre.addWidget(btn_ajouter); barre.addWidget(btn_supprimer); barre.addWidget(btn_importer)
            barre.addStretch(1)
            layout.addLayout(barre)

        self.table = QTableWidget(0, 4)
        self.table.setHorizontalHeaderLabels(["Zone", "Chapelle", "Planche", "Dernière culture"])
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        layout.addWidget(self.table, 1)

        if mode == "zero":
            self._generer_defaut()

        boutons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        boutons.accepted.connect(self._valider)
        boutons.rejected.connect(self.reject)
        layout.addWidget(boutons)

    # -- mode "zero" ----------------------------------------------------
    def _generer_defaut(self):
        planches = generer_disposition_defaut(
            self.spin_zones.value(), self.spin_chapelles.value(), self.spin_planches.value())
        self._remplir_table(planches, editable_culture=False)

    # -- mode "existant" -------------------------------------------------
    def _ajouter_ligne_vide(self):
        row = self.table.rowCount()
        self.table.insertRow(row)
        self.table.setItem(row, 0, QTableWidgetItem("Zone 1"))
        self.table.setItem(row, 1, QTableWidgetItem("Chapelle 1"))
        self.table.setItem(row, 2, QTableWidgetItem(f"Planche {row + 1}"))
        self._poser_combo_culture(row, "")

    def _supprimer_ligne(self):
        rows = sorted({i.row() for i in self.table.selectedIndexes()}, reverse=True)
        for r in rows:
            self.table.removeRow(r)

    def _importer_plan(self):
        chemin, _ = QFileDialog.getOpenFileName(
            self, "Importer un plan de ferme", "", "Plans de ferme (*.ferme *.txt);;Tous les fichiers (*.*)")
        if not chemin:
            return
        try:
            planches = importer_planches_depuis_plan(chemin, self.pf)
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Impossible de lire le plan de ferme :\n{e}")
            return
        if not planches:
            QMessageBox.information(self, "Info", "Aucune planche trouvée dans ce plan de ferme.")
            return
        self._remplir_table(planches, editable_culture=True)

    def _poser_combo_culture(self, row, culture_actuelle):
        combo = QComboBox()
        combo.addItem("(aucune)")
        combo.addItems(sorted(self.fb.FICHES.keys()))
        if culture_actuelle and culture_actuelle in self.fb.FICHES:
            combo.setCurrentText(culture_actuelle)
        else:
            combo.setCurrentIndex(0)
        self.table.setCellWidget(row, 3, combo)

    def _remplir_table(self, planches, editable_culture):
        self.table.setRowCount(0)
        for p in planches:
            row = self.table.rowCount()
            self.table.insertRow(row)
            self.table.setItem(row, 0, QTableWidgetItem(p["zone"]))
            self.table.setItem(row, 1, QTableWidgetItem(p["chapelle"]))
            self.table.setItem(row, 2, QTableWidgetItem(p["planche"]))
            if self.mode == "zero":
                item_culture = QTableWidgetItem("—")
                item_culture.setFlags(item_culture.flags() & ~Qt.ItemIsEditable)
                self.table.setItem(row, 3, item_culture)
            else:
                self._poser_combo_culture(row, p.get("derniere_culture", ""))

    def _valider(self):
        if self.table.rowCount() == 0:
            QMessageBox.warning(self, "Disposition vide", "Ajoutez au moins une planche.")
            return
        self.accept()

    def planches(self):
        resultats = []
        for row in range(self.table.rowCount()):
            zone = self.table.item(row, 0).text().strip()
            chapelle = self.table.item(row, 1).text().strip()
            planche = self.table.item(row, 2).text().strip()
            if self.mode == "zero":
                derniere = ""
            else:
                combo = self.table.cellWidget(row, 3)
                texte = combo.currentText() if combo else ""
                derniere = "" if texte == "(aucune)" else texte
            resultats.append({"zone": zone, "chapelle": chapelle, "planche": planche,
                               "derniere_culture": derniere})
        return resultats


class DialogueSelectionCultures(QDialog):
    """Pop-up 5 : pour chaque planche, choix d'une culture parmi celles
    triées par score d'adéquation décroissant."""

    def __init__(self, parent, planches, nom_region, semaine_ref, rows_planning, fb, rc, cf):
        super().__init__(parent)
        self.planches = planches
        self.nom_region = nom_region
        self.semaine_ref = semaine_ref
        self.rows_planning = rows_planning
        self.fb, self.rc, self.cf = fb, rc, cf
        self.assignations = {}   # index planche -> nom culture
        self.scores = {}         # index planche -> score

        self.setWindowTitle("Créer un planning cultural — Sélection des cultures")
        self.resize(920, 560)
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(
            "Sélectionnez une planche à gauche, puis assignez-lui une culture dans la liste "
            "de droite (triée par score d'adéquation région + saison"
            + (" + rotation" if any(p["derniere_culture"] for p in planches) else "")
            + "). Les cultures grisées violent la règle de rotation (même famille "
            "botanique que la dernière culture de la planche)."))

        splitter = QSplitter(Qt.Horizontal)
        layout.addWidget(splitter, 1)

        self.liste_planches = QListWidget()
        for p in planches:
            self.liste_planches.addItem(self._libelle_planche(p))
        self.liste_planches.currentRowChanged.connect(self._changer_planche)
        splitter.addWidget(self.liste_planches)

        self.table_cultures = QTableWidget(0, 4)
        self.table_cultures.setHorizontalHeaderLabels(["Culture", "Score", "Famille", "Fenêtre / motif"])
        self.table_cultures.horizontalHeader().setStretchLastSection(True)
        self.table_cultures.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table_cultures.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table_cultures.itemDoubleClicked.connect(self._assigner_depuis_table)
        splitter.addWidget(self.table_cultures)
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 2)

        barre = QHBoxLayout()
        btn_assigner = QPushButton("✅ Assigner la culture sélectionnée à cette planche")
        btn_assigner.clicked.connect(self._assigner_depuis_table)
        barre.addWidget(btn_assigner)
        btn_retirer = QPushButton("✖ Retirer l'assignation de cette planche")
        btn_retirer.clicked.connect(self._retirer_assignation)
        barre.addWidget(btn_retirer)
        barre.addStretch(1)
        layout.addLayout(barre)

        boutons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        boutons.accepted.connect(self._valider)
        boutons.rejected.connect(self.reject)
        layout.addWidget(boutons)

        if planches:
            self.liste_planches.setCurrentRow(0)

    def _libelle_planche(self, p):
        base = f"{p['zone']} / {p['chapelle']} / {p['planche']}"
        if p["derniere_culture"]:
            base += f"  (préc. : {p['derniere_culture']})"
        return base

    def _rafraichir_libelle(self, idx):
        item = self.liste_planches.item(idx)
        p = self.planches[idx]
        texte = self._libelle_planche(p)
        if idx in self.assignations:
            texte += f"  ➜ {self.assignations[idx]} ({self.scores[idx]}/100)"
        item.setText(texte)

    def _changer_planche(self, idx):
        if idx < 0:
            return
        p = self.planches[idx]
        cultures = lister_cultures_triees(
            self.nom_region, self.semaine_ref, p["derniere_culture"],
            self.rows_planning, self.fb, self.rc, self.cf)

        self.table_cultures.setRowCount(len(cultures))
        for row, c in enumerate(cultures):
            item_nom = QTableWidgetItem(c["culture"])
            item_score = QTableWidgetItem(f"{c['score']}/100")
            item_famille = QTableWidgetItem(c["famille"])
            motif = c["motif_rotation"] or c["fenetre"]
            item_motif = QTableWidgetItem(motif)
            if c["interdit"]:
                for it in (item_nom, item_score, item_famille, item_motif):
                    it.setFlags(it.flags() & ~Qt.ItemIsEnabled)
                    it.setBackground(QBrush(QColor("#f0f0f0")))
                    it.setForeground(QBrush(QColor("#a0a0a0")))
                item_motif.setText(f"❌ interdit : {c['motif_rotation']}")
            elif c["score"] >= 70:
                item_score.setForeground(QBrush(QColor("#2e7d32")))
            self.table_cultures.setItem(row, 0, item_nom)
            self.table_cultures.setItem(row, 1, item_score)
            self.table_cultures.setItem(row, 2, item_famille)
            self.table_cultures.setItem(row, 3, item_motif)

    def _assigner_depuis_table(self, *_args):
        idx_planche = self.liste_planches.currentRow()
        row = self.table_cultures.currentRow()
        if idx_planche < 0 or row < 0:
            return
        item = self.table_cultures.item(row, 0)
        if not (item.flags() & Qt.ItemIsEnabled):
            QMessageBox.warning(self, "Rotation non respectée",
                                 "Cette culture ne respecte pas la règle de rotation pour "
                                 "cette planche et ne peut pas être sélectionnée.")
            return
        culture = item.text()
        score = int(self.table_cultures.item(row, 1).text().split("/")[0])
        self.assignations[idx_planche] = culture
        self.scores[idx_planche] = score
        self._rafraichir_libelle(idx_planche)

    def _retirer_assignation(self):
        idx_planche = self.liste_planches.currentRow()
        if idx_planche < 0:
            return
        self.assignations.pop(idx_planche, None)
        self.scores.pop(idx_planche, None)
        self._rafraichir_libelle(idx_planche)

    def _valider(self):
        if not self.assignations:
            QMessageBox.warning(self, "Aucune sélection",
                                 "Assignez une culture à au moins une planche.")
            return
        self.accept()


# ------------------------------------------------------------------
# Orchestration de l'assistant
# ------------------------------------------------------------------
def lancer_assistant(app, fb, rc, pf, cf):
    """Lance les pop-ups de l'assistant dans l'ordre. Renvoie
    (planches, assignations, scores, semaine_debut_periode,
    semaine_fin_periode) ou None si l'utilisateur a annulé à une étape."""

    dlg1 = DialoguePointDepart(app)
    if dlg1.exec_() != QDialog.Accepted:
        return None
    mode = dlg1.mode()

    dlg2 = DialogueDureeCulture(app)
    if dlg2.exec_() != QDialog.Accepted:
        return None
    semaine_debut, semaine_fin = dlg2.periode()

    dlg3 = DialogueChoixRegion(app, cf)
    if dlg3.exec_() != QDialog.Accepted:
        return None
    region = dlg3.region_choisie()

    dlg4 = DialogueDispositionFerme(app, mode, fb, pf)
    if dlg4.exec_() != QDialog.Accepted:
        return None
    planches = dlg4.planches()

    semaine_ref = semaine_debut
    dlg5 = DialogueSelectionCultures(app, planches, region, semaine_ref, app.rows, fb, rc, cf)
    if dlg5.exec_() != QDialog.Accepted:
        return None

    return planches, dlg5.assignations, dlg5.scores, semaine_debut, semaine_fin


# ====================================================================
# Construction de l'onglet
# ====================================================================
def construire_onglet_creation_planning(app, parent, fb, rc, pf, cf):
    """Construit l'onglet « Création planning cultural » et attache les
    gestionnaires d'évènements sur `app` (préfixe _ocp_)."""
    layout = QVBoxLayout(parent)

    intro = QLabel(
        "Cet assistant vous guide pas à pas pour construire un planning cultural : "
        "point de départ, période, région, disposition de la ferme, puis sélection des "
        "cultures les mieux adaptées (triées par score). Le résultat est rassemblé dans "
        "le tableau ci-dessous, que vous pouvez ajouter au planning cultural ou exporter "
        "vers le fichier Excel."
    )
    intro.setWordWrap(True)
    intro.setStyleSheet("color:#555; font-style:italic; font-size:9pt;")
    layout.addWidget(intro)

    barre_haut = QHBoxLayout()
    app._ocp_btn_lancer = QPushButton("🌱 Créer un planning cultural")
    app._ocp_btn_lancer.setStyleSheet("font-weight:bold; padding:8px; font-size:11pt;")
    app._ocp_btn_lancer.clicked.connect(lambda: _lancer(app, fb, rc, pf, cf))
    barre_haut.addWidget(app._ocp_btn_lancer)
    barre_haut.addStretch(1)
    layout.addLayout(barre_haut)

    app._ocp_resume = QLabel("")
    app._ocp_resume.setWordWrap(True)
    layout.addWidget(app._ocp_resume)

    app._ocp_table = QTableWidget(0, len(ENTETES_FINALES))
    app._ocp_table.setHorizontalHeaderLabels(ENTETES_FINALES)
    app._ocp_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
    app._ocp_table.setAlternatingRowColors(True)
    app._ocp_table.horizontalHeader().setStretchLastSection(True)
    app._ocp_table.verticalHeader().setVisible(False)
    layout.addWidget(app._ocp_table, 1)

    barre_bas = QHBoxLayout()
    app._ocp_btn_ajouter = QPushButton("➕ Ajouter ces lignes au planning cultural")
    app._ocp_btn_ajouter.clicked.connect(lambda: _ajouter_au_planning(app))
    app._ocp_btn_ajouter.setEnabled(False)
    barre_bas.addWidget(app._ocp_btn_ajouter)

    app._ocp_btn_excel = QPushButton("📊 Exporter vers le fichier Excel...")
    app._ocp_btn_excel.clicked.connect(lambda: _exporter_excel(app))
    app._ocp_btn_excel.setEnabled(False)
    barre_bas.addWidget(app._ocp_btn_excel)
    barre_bas.addStretch(1)
    layout.addLayout(barre_bas)

    app._ocp_lignes = []


def _lancer(app, fb, rc, pf, cf):
    resultat = lancer_assistant(app, fb, rc, pf, cf)
    if resultat is None:
        return
    planches, assignations, scores, sem_debut, sem_fin = resultat
    lignes = construire_lignes_finales(planches, assignations, scores, sem_debut, sem_fin, app.rows)
    app._ocp_lignes = lignes

    app._ocp_table.setRowCount(len(lignes))
    for row, ligne in enumerate(lignes):
        for col, champ in enumerate(COLONNES_FINALES):
            app._ocp_table.setItem(row, col, QTableWidgetItem(str(ligne.get(champ, ""))))

    nb_planches_assignees = len(assignations)
    app._ocp_resume.setText(
        f"{nb_planches_assignees} planche(s) sur {len(planches)} assignée(s) — "
        f"{len(lignes)} ligne(s) générée(s) pour les semaines {sem_debut} à {sem_fin}."
    )
    app._ocp_btn_ajouter.setEnabled(bool(lignes))
    app._ocp_btn_excel.setEnabled(bool(lignes))
    app._set_status(f"Planning créé : {len(lignes)} ligne(s).")


def _ajouter_au_planning(app):
    if not app._ocp_lignes:
        return
    ajoutees = 0
    for ligne in app._ocp_lignes:
        if not ligne.get("action") or ligne["action"] == "(aucune donnée de planning)":
            continue
        commentaire = ligne.get("commentaire", "")
        prefixe = f"[{ligne['zone']} / {ligne['chapelle']} / {ligne['planche']}]"
        commentaire = f"{prefixe} {commentaire}".strip()
        app.rows.append({
            "culture": ligne["culture"], "conduite": ligne.get("conduite", ""),
            "variete_n": ligne.get("variete_n", "") or "1", "action": ligne["action"],
            "semaine_debut": str(ligne.get("semaine_debut", "")),
            "semaine_fin": str(ligne.get("semaine_fin", "")),
            "mois_debut": ligne.get("mois_debut", ""), "mois_fin": ligne.get("mois_fin", ""),
            "commentaire": commentaire,
        })
        ajoutees += 1
    app._rafraichir_tableau()
    if hasattr(app, "table_semaine"):
        app._rafraichir_actions_semaine()
    if hasattr(app, "cb_dec_culture"):
        app._maj_liste_cultures()
    if hasattr(app, "liste_fiches"):
        app._maj_liste_fiches()
    QMessageBox.information(app, "Ajouté", f"{ajoutees} ligne(s) ajoutée(s) au planning cultural.\n"
                             "Pensez à enregistrer (Fichier → Enregistrer) pour conserver le CSV.")


def _exporter_excel(app):
    if not app._ocp_lignes:
        return
    chemin_defaut = getattr(app, "csv_path", "") or ""
    chemin_defaut = os.path.splitext(chemin_defaut)[0] + ".xlsx" if chemin_defaut else "planning_cultural.xlsx"
    chemin, _ = QFileDialog.getSaveFileName(
        app, "Exporter vers le fichier Excel", chemin_defaut, "Classeurs Excel (*.xlsx)")
    if not chemin:
        return
    if not chemin.lower().endswith(".xlsx"):
        chemin += ".xlsx"
    try:
        exporter_vers_excel(chemin, app._ocp_lignes)
    except ImportError:
        QMessageBox.critical(app, "Erreur", "Le paquet 'openpyxl' est requis.\nInstallez-le avec : pip install openpyxl")
        return
    except Exception as e:
        QMessageBox.critical(app, "Erreur", f"Impossible d'exporter vers Excel :\n{e}")
        return
    app._set_status(f"Tableau exporté vers {chemin} (feuille « Planning créé (assistant) »).")
    QMessageBox.information(app, "Export réussi", f"Le tableau a été ajouté au fichier :\n{chemin}")
